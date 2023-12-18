from openpyxl import *
from tkinter import messagebox
import os
import restaurant_additional_details as r

def data(direct, registration_number, email_id, password_, restaurant_name, phone_number, address_, state_, pincode_):
    column_names = {"Registration Number": "A", "Email": "B", "Password": "C", "Name": "D", "Phone Number": "E",
                    "Address": "F", "State": "G", "Pincode": "H"}

    try:
        excelloc = os.path.join(direct, r'User_Data.xlsx')
        workbook = load_workbook(excelloc)
        worksheet = workbook["User Data"]

    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "User Data"
        worksheet.append(list(column_names.keys()))
        workbook.save(excelloc)
        
    else:
        worksheet[f'{column_names["Registration Number"]}{worksheet.max_row+1}'] = registration_number
        worksheet[f'{column_names["Email"]}{worksheet.max_row}'] = email_id
        worksheet[f'{column_names["Password"]}{worksheet.max_row}'] = password_
        worksheet[f'{column_names["Name"]}{worksheet.max_row}'] = restaurant_name
        worksheet[f'{column_names["Phone Number"]}{worksheet.max_row}'] = phone_number
        worksheet[f'{column_names["Address"]}{worksheet.max_row}'] = address_
        worksheet[f'{column_names["State"]}{worksheet.max_row}'] = state_
        worksheet[f'{column_names["Pincode"]}{worksheet.max_row}'] = pincode_
        workbook.save(excelloc)

    workbook.close()
    r.working_hours(direct, registration_number)
