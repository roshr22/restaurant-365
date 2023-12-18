from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import customtkinter
from PIL import ImageTk
import os
from openpyxl import *

def staffdetailspage(direct, datadirect):
    def landingpg(root, direct, datadirect):
        root.destroy()
        import landingpage
        landingpage.landingpage(direct, datadirect)
    
    def addstaff(direct, datadirect):
        def sid(event):
            if staffid.get() == 'Enter Staff Id':
                staffid.delete(0, END)
                
        def sname(event):
            if staffname.get() == 'Enter Staff Name':
                staffname.delete(0, END)

        def pos(event):
            if position.get() == 'Enter Position':
                position.delete(0, END)

        def basesal(event):
            if basesalary.get() == 'Enter Base Salary':
                basesalary.delete(0, END)

        def bon(event):
            if bonus.get() == 'Enter Bonus':
                bonus.delete(0, END)

        def adddata(datadirect, sid, name, position, base, bonus):
            workbook = load_workbook(datadirect + r"\Staff Management.xlsx")
            worksheet = workbook["Staff Details"]
            column_names = {"Employee ID":"A","Employee Name":"B", "Position": "C", "Base Salary":"D", "Bonus":"E", "Total Salary":"F"}

            if not all(x.isalpha() or x.isdigit() for x in sid):
                messagebox.showerror("Error", f"Invalid staff id")
                return
            
            if not all(x.isalpha() or x.isspace() for x in name):
                messagebox.showerror("Error", f"Invalid staff name")
                return

            if not all(x.isalpha() or x.isspace() for x in position):
                messagebox.showerror("Error", f"Invalid staff position")
                return

            if not (base.isdigit() or ((all(x.isdigit() or x == "." for x in base)) and base.count(".") <= 1)):
                messagebox.showerror("Error", f"Invalid base salary")
                return

            if not (bonus.isdigit() or ((all(x.isdigit() or x == "." for x in bonus)) and bonus.count(".") <= 1)):
                messagebox.showerror("Error", f"Invalid bonus")
                return
            
            for i in range(2, worksheet.max_row + 1):
                sid_excel = worksheet[f'{column_names["Employee ID"]}{i}'].value
                name_excel = worksheet[f'{column_names["Employee Name"]}{i}'].value
                position_excel = worksheet[f'{column_names["Position"]}{i}'].value
                base_excel = worksheet[f'{column_names["Base Salary"]}{i}'].value
                bonus_excel = worksheet[f'{column_names["Bonus"]}{i}'].value
                if sid_excel.lower() == sid.lower() and name_excel.lower() == name.lower():
                    messagebox.showerror("Error", f"Staff Exists")
                    break
                elif sid_excel.lower() == sid.lower():
                    messagebox.showerror("Error", f"Staff with same id exists... Try a different id")
                    break
            else:
                worksheet[f'{column_names["Employee ID"]}{worksheet.max_row+1}'] = sid
                worksheet[f'{column_names["Employee Name"]}{worksheet.max_row}'] = name.title()
                worksheet[f'{column_names["Position"]}{worksheet.max_row}'] = position.title()
                worksheet[f'{column_names["Base Salary"]}{worksheet.max_row}'] = float(base)
                worksheet[f'{column_names["Bonus"]}{worksheet.max_row}'] = float(bonus)
                worksheet[f'{column_names["Total Salary"]}{worksheet.max_row}'] = float(base) + float(bonus)
                workbook.save(datadirect + r"\Staff Management.xlsx")
                messagebox.showinfo(message=f"Staff details have been added to database")
        
        customtkinter.set_appearance_mode("dark")
        root = Tk()
        root.geometry('1366x768')
        root.state('zoomed')

        nsploc = os.path.join(direct, r'images\newstaffpage.png')
        bgimage = ImageTk.PhotoImage(file=nsploc)
        bglabel = Label(root, image=bgimage)
        bglabel.place(x=0, y=0)

        # Staff Id
        staffidlabel = Label(root, bg="white", fg="black", text="Staff Id", font=('Garamond', 23), bd=0)
        staffidlabel.place(x=460, y=150)
        staffid = Entry(root, bg="white", fg="grey", font=('Garamond', 20), width=28, bd=0)
        staffid.place(x=460, y=190)
        staffid.insert(0, "Enter Staff Id")
        staffid.bind('<FocusIn>', sid)
        line3 = Frame(root, height=2, width=375, bg="black")
        line3.place(x=455, y=220)


        # Staff Name
        staffnamelabel = Label(root, bg="white", fg="black", text="Staff Name", font=('Garamond', 23), bd=0)
        staffnamelabel.place(x=460, y=240)
        staffname = Entry(root, bg="white", fg="grey", font=('Garamond', 20), width=28, bd=0)
        staffname.place(x=460, y=280)
        staffname.insert(0, "Enter Staff Name")
        staffname.bind('<FocusIn>', sname)
        line1 = Frame(root, height=2, width=375, bg="black")
        line1.place(x=455, y=310)

        #Position
        positionlabel = Label(root, bg="white", fg="black", text="Position", font=('Garamond', 23), bd=0)
        positionlabel.place(x=460, y=330)
        position = Entry(root, bg="white", fg="grey", font=('Garamond', 20), width=28, bd=0)
        position.place(x=460, y=370)
        position.insert(0, "Enter Position")
        position.bind('<FocusIn>', pos)
        line5 = Frame(root, height=2, width=375, bg="black")
        line5.place(x=455, y=400)

        # Base Salary
        basesalarylabel = Label(root, bg="white", fg="black", text="Base Salary", font=('Garamond', 23), bd=0)
        basesalarylabel.place(x=460, y=420)
        basesalary = Entry(root, bg="white", fg="grey", font=('Garamond', 20), width=28, bd=0)
        basesalary.place(x=460, y=460)
        basesalary.insert(0, "Enter Base Salary")
        basesalary.bind('<FocusIn>', basesal)
        line4 = Frame(root, height=2, width=375, bg="black")
        line4.place(x=455, y=490)

        # Bonus
        bonuslabel = Label(root, bg="white", fg="black", text="Bonus", font=('Garamond', 23), bd=0)
        bonuslabel.place(x=460, y=510)
        bonus = Entry(root, bg="white", fg="grey", font=('Garamond', 20), width=28, bd=0)
        bonus.place(x=460, y=550)
        bonus.insert(0, "Enter Bonus")
        bonus.bind('<FocusIn>', bon)
        line5 = Frame(root, height=2, width=375, bg="black")
        line5.place(x=455, y=580)

        addbutton = Button(root, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                                     text="Add", font=('Garamond', 20), width=5, bd=0, command = lambda:(adddata(datadirect, staffid.get().strip(), staffname.get().strip(), position.get().strip(), basesalary.get().strip(), bonus.get().strip())), cursor='hand2')
        addbutton.place(x=525, y=590)

        donebutton = Button(root, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                                     text="Done", font=('Garamond', 20), width=5, bd=0, command = lambda:(root.destroy(), staffdetailspage(direct, datadirect)), cursor='hand2')
        donebutton.place(x=690, y=590)

        root.mainloop()

    def modifydetails(direct, datadirect):
        def modifyexcel(datadirect, root, positions, basesalaries, bonuses):
            column_names = {"Employee ID":"A","Employee Name":"B", "Position": "C", "Base Salary":"D", "Bonus":"E", "Total Salary":"F"}
            workbook = load_workbook(datadirect + r"\Staff Management.xlsx")
            worksheet = workbook["Staff Details"]

            for i in range(2, worksheet.max_row + 1):
                position = positions[i-1].get().strip()
                base = basesalaries[i-1].get().strip()
                bonus = bonuses[i-1].get().strip()

                if not all(x.isalpha() or x.isspace() for x in position):
                    messagebox.showerror("Error", f"Invalid staff position")
                    return

                if not (base.isdigit() or ((all(x.isdigit() or x == "." for x in base)) and base.count(".") <= 1)):
                    messagebox.showerror("Error", f"Invalid base salary")
                    return

                if not (bonus.isdigit() or ((all(x.isdigit() or x == "." for x in bonus)) and bonus.count(".") <= 1)):
                    messagebox.showerror("Error", f"Invalid bonus")
                    return

                worksheet[f'{column_names["Position"]}{worksheet.max_row}'] = position.title()
                worksheet[f'{column_names["Base Salary"]}{worksheet.max_row}'] = float(base)
                worksheet[f'{column_names["Bonus"]}{worksheet.max_row}'] = float(bonus)
                worksheet[f'{column_names["Total Salary"]}{worksheet.max_row}'] = float(base) + float(bonus)
                workbook.save(datadirect + r"\Staff Management.xlsx")

            root.destroy()
            staffdetailspage(direct, datadirect)
                
        
        column = {"Employee ID":"A","Employee Name":"B", "Position": "C", "Base Salary":"D", "Bonus":"E", "Total Salary":"F"}
        workbook = load_workbook(datadirect + r"\Staff Management.xlsx")
        worksheet = workbook["Staff Details"]
        
        customtkinter.set_appearance_mode("dark")
        root1 = Tk()
        root1.geometry('1366x768')
        root1.state('zoomed')

        msdloc = os.path.join(direct, r'images\modifystaffdetails.png')
        bgimage1 = ImageTk.PhotoImage(file=msdloc)
        bglabel = Label(root1, image=bgimage1)
        bglabel.place(x=0, y=0)

        # Create the canvas widget
        canvas = Canvas(root1, width=840, height=355, bg="white", bd=0, highlightbackground="white", highlightcolor="white")

        # Create the scrollbar widget
        scroll = Scrollbar(root1, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
        canvas.config(yscrollcommand=scroll.set)

        # Create the frame widget
        frame = Frame(canvas, width=830, height=345, bg="white")

        positions = []
        basesalaries = []
        bonuses = []
        # Create labels and entry boxes and add them to the frame
        for j in range(1, int(worksheet.max_row) + 1):
            employeeidvalue = worksheet[f'{column["Employee ID"]}{j}'].value
            employeenamevalue = worksheet[f'{column["Employee Name"]}{j}'].value
            positionvalue = worksheet[f'{column["Position"]}{j}'].value
            basesalaryvalue = worksheet[f'{column["Base Salary"]}{j}'].value
            bonusvalue = worksheet[f'{column["Bonus"]}{j}'].value
            exec(f'employeeid{j}entry = Entry(frame, bg="white", fg="black", readonlybackground="white", font=("Garamond", 18), bd=0, width=15)')
            exec(f'employeename{j}entry = Entry(frame, bg="white", fg="black", readonlybackground="white", font=("Garamond", 18), bd=0, width=20)')
            exec(f'position{j}entry = Entry(frame, bg="white", fg="black", readonlybackground="white", font=("Garamond", 18), bd=0, width=15)')
            exec(f'basesalary{j}entry = Entry(frame, bg="white", fg="black", readonlybackground="white", font=("Garamond", 18), bd=0, width=10)')
            exec(f'bonus{j}entry = Entry(frame, bg="white", fg="black", readonlybackground="white", font=("Garamond", 18), bd=0, width=10)')
            exec(f'employeeid{j}entry.insert(0, employeeidvalue)')
            exec(f'employeename{j}entry.insert(0, employeenamevalue)')
            exec(f'position{j}entry.insert(0, str(positionvalue))')
            exec(f'basesalary{j}entry.insert(0, str(basesalaryvalue))')
            exec(f'bonus{j}entry.insert(0, str(bonusvalue))')
            exec(f'employeeid{j}entry["state"] = "readonly"')
            exec(f'employeename{j}entry["state"] = "readonly"')
            if j == 1:
                exec(f'position{j}entry["state"] = "readonly"')
                exec(f'basesalary{j}entry["state"] = "readonly"')
                exec(f'bonus{j}entry["state"] = "readonly"')
            if j != 1:
                exec(f'position{j}entry.configure(highlightthickness=1, highlightbackground="black", highlightcolor="black")')
                exec(f'basesalary{j}entry.configure(highlightthickness=1, highlightbackground="black", highlightcolor="black")')
                exec(f'bonus{j}entry.configure(highlightthickness=1, highlightbackground="black", highlightcolor="black")')
            exec(f'employeeid{j}entry.grid(row=j, column=0, padx=0, pady=10)')
            exec(f'employeename{j}entry.grid(row=j, column=1, padx=0, pady=10)')
            exec(f'position{j}entry.grid(row=j, column=2, padx=0, pady=10)')
            exec(f'basesalary{j}entry.grid(row=j, column=3, padx=15, pady=10)')
            exec(f'bonus{j}entry.grid(row=j, column=4, padx=0, pady=10)')
            exec(f'positions.append(position{j}entry)')
            exec(f'basesalaries.append(basesalary{j}entry)')
            exec(f'bonuses.append(bonus{j}entry)')

        # Add the frame to the canvas
        canvas.create_window(10, 10, window=frame, anchor="nw")

        # Update the scrollregion of the canvas
        root1.update()
        canvas.configure(scrollregion = canvas.bbox('all'))

        # Place the canvas and the scrollbar in the window
        canvas.place(x=200, y=200)
        scroll.place(x=1040, y=200, height=355)

        modifydetailsbutton = Button(root1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="Modify Details", font=('Garamond', 20), bd=0, command = lambda:(modifyexcel(datadirect, root1, positions, basesalaries, bonuses)), cursor='hand2')
        modifydetailsbutton.place(x=545, y=570)
    
        root1.mainloop()

    def removestaff(direct, datadirect):
        global removes
        global root1
        def removes(root, i):
            column_names = {"Employee ID":"A","Employee Name":"B", "Position": "C", "Base Salary":"D", "Bonus":"E", "Total Salary":"F"}
            workbook = load_workbook(datadirect + r"\Staff Management.xlsx")
            worksheet = workbook["Staff Details"]

            for j in range(i, int(worksheet.max_row)):
                worksheet[f'{column_names["Employee ID"]}{j}'] = worksheet[f'{column_names["Employee ID"]}{j+1}']
                worksheet[f'{column_names["Employee Name"]}{j}'] = worksheet[f'{column_names["Employee Name"]}{j+1}']
                worksheet[f'{column_names["Position"]}{j}'] = worksheet[f'{column_names["Position"]}{j+1}']
                worksheet[f'{column_names["Base Salary"]}{j}'] = worksheet[f'{column_names["Base Salary"]}{j+1}']
                worksheet[f'{column_names["Bonus"]}{j}'] = worksheet[f'{column_names["Bonus"]}{j+1}']
                worksheet[f'{column_names["Total Salary"]}{j}'] = worksheet[f'{column_names["Total Salary"]}{j+1}']
                workbook.save(datadirect + r"\Staff Management.xlsx")
                
            else:
                worksheet[f'{column_names["Employee ID"]}{worksheet.max_row}'] = None
                worksheet[f'{column_names["Employee Name"]}{worksheet.max_row}'] = None
                worksheet[f'{column_names["Position"]}{worksheet.max_row}'] = None
                worksheet[f'{column_names["Base Salary"]}{worksheet.max_row}'] = None
                worksheet[f'{column_names["Bonus"]}{worksheet.max_row}'] = None
                worksheet[f'{column_names["Total Salary"]}{worksheet.max_row}'] = None
                workbook.save(datadirect + r"\Staff Management.xlsx")
                
            root.destroy()
            staffdetailspage(direct, datadirect)
            
        column = {"Employee ID":"A","Employee Name":"B", "Position": "C", "Base Salary":"D", "Bonus":"E", "Total Salary":"F"}
        workbook = load_workbook(datadirect + r"\Staff Management.xlsx")
        worksheet = workbook["Staff Details"]
        
        customtkinter.set_appearance_mode("dark")
        root1 = Tk()
        root1.geometry('1366x768')
        root1.state('zoomed')

        rsdimg = os.path.join(direct, r'images\removestaffdetails.png')
        bgimage1 = ImageTk.PhotoImage(file=rsdimg)
        bglabel = Label(root1, image=bgimage1)
        bglabel.place(x=0, y=0)

        # Create the canvas widget
        canvas = Canvas(root1, width=600, height=355, bg="white", bd=0, highlightbackground="white", highlightcolor="white")

        # Create the scrollbar widget
        scroll = Scrollbar(root1, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
        canvas.config(yscrollcommand=scroll.set)

        # Create the frame widget
        frame = Frame(canvas, width=590, height=345, bg="white")

        # Create labels and entry boxes and add them to the frame
        for j in range(1, int(worksheet.max_row) + 1):
            employeeidvalue = worksheet[f'{column["Employee ID"]}{j}'].value
            employeenamevalue = worksheet[f'{column["Employee Name"]}{j}'].value
            positionvalue = worksheet[f'{column["Position"]}{j}'].value
            exec(f'employeeid{j}entry = Label(frame, text=employeeidvalue, bg="white", fg="black", font=("Garamond", 18), bd=0)')
            exec(f'employeename{j}entry = Label(frame, text=employeenamevalue, bg="white", fg="black", font=("Garamond", 18), bd=0)')
            exec(f'position{j}entry = Label(frame, text=positionvalue, bg="white", fg="black", font=("Garamond", 18), bd=0)')
            exec(f'employeeid{j}entry.grid(row=j, column=0, padx=15, pady=10)')
            exec(f'employeename{j}entry.grid(row=j, column=1, padx=15, pady=10)')
            exec(f'position{j}entry.grid(row=j, column=2, padx=15, pady=10)')
            if j != 1:
                exec(f'remove{j}button = Button(frame, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="Remove", font=("Garamond", 18), bd=0, cursor="hand2", width=8, command=lambda:(removes(root1, {j})))')
                exec(f'remove{j}button.grid(row=j, column=3, padx=15, pady=10)')                

        # Add the frame to the canvas
        canvas.create_window(10, 10, window=frame, anchor="nw")

        # Update the scrollregion of the canvas
        root1.update()
        canvas.configure(scrollregion = canvas.bbox('all'))

        # Place the canvas and the scrollbar in the window
        canvas.place(x=330, y=200)
        scroll.place(x=930, y=200, height=355)

        donebutton = Button(root1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                                     text="Done", font=('Garamond', 20), width=5, bd=0, command = lambda:(root1.destroy(), staffdetailspage(direct, datadirect)), cursor='hand2')
        donebutton.place(x=550, y=590)


        root1.mainloop()

    column = {"Employee ID":"A","Employee Name":"B", "Position": "C", "Base Salary":"D", "Bonus":"E", "Total Salary":"F"}
    try:
        workbook = load_workbook(datadirect + r"\Staff Management.xlsx")
        worksheet = workbook["Staff Details"]

    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Staff Details"
        worksheet.append(list(column.keys()))
        workbook.save(datadirect + r"\Staff Management.xlsx")
    
    customtkinter.set_appearance_mode("dark")
    root1 = Tk()
    root1.geometry('1366x768')
    root1.state('zoomed')

    smploc = os.path.join(direct, r'images\staffmanagementpage.png')
    bgimage1 = ImageTk.PhotoImage(file=smploc)
    bglabel = Label(root1, image=bgimage1)
    bglabel.place(x=0, y=0)

    liloc = os.path.join(direct, r'images\logohomebutton.png')
    logoimage = ImageTk.PhotoImage(file=liloc)
    homebutton = Button(root1, image = logoimage , fg='white', bg='white', activeforeground='white',activebackground='white', cursor='hand2', bd=0, command=lambda:(landingpg(root1, direct, datadirect)))
    homebutton.place(x=20, y=20)

    # Create the canvas widget
    canvas = Canvas(root1, width=840, height=355, bg="white", bd=0, highlightbackground="white", highlightcolor="white")

    # Create the scrollbar widget
    scroll = Scrollbar(root1, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
    canvas.config(yscrollcommand=scroll.set)

    # Create the frame widget
    frame = Frame(canvas, width=830, height=345, bg="white")

    capacity = []
    quantity = []
    # Create labels and entry boxes and add them to the frame
    for j in range(1, int(worksheet.max_row) + 1):
        employeeidvalue = worksheet[f'{column["Employee ID"]}{j}'].value
        employeenamevalue = worksheet[f'{column["Employee Name"]}{j}'].value
        positionvalue = worksheet[f'{column["Position"]}{j}'].value
        basesalaryvalue = worksheet[f'{column["Base Salary"]}{j}'].value
        bonusvalue = worksheet[f'{column["Bonus"]}{j}'].value
        totalsalaryvalue = worksheet[f'{column["Total Salary"]}{j}'].value
        exec(f'employeeid{j}label = Label(frame, text=employeeidvalue, bg="white", fg="black", font=("Garamond", 18), bd=0)')
        exec(f'employeename{j}label = Label(frame, text=employeenamevalue, bg="white", fg="black", wraplength=170, font=("Garamond", 18), bd=0)')
        exec(f'position{j}label = Label(frame, text=positionvalue, bg="white", fg="black",wraplength=170, font=("Garamond", 18), bd=0)')
        exec(f'basesalary{j}label = Label(frame, text=basesalaryvalue, bg="white", fg="black", font=("Garamond", 18), bd=0)')
        exec(f'bonus{j}label = Label(frame, text=bonusvalue, bg="white", fg="black", font=("Garamond", 18), bd=0)')
        exec(f'totalsalary{j}label = Label(frame, text=totalsalaryvalue, bg="white", fg="black", font=("Garamond", 18), bd=0)')
        exec(f'employeeid{j}label.grid(row=j, column=0, padx=15, pady=10)')
        exec(f'employeename{j}label.grid(row=j, column=1, padx=15, pady=10)')
        exec(f'position{j}label.grid(row=j, column=2, padx=15, pady=10)')
        exec(f'basesalary{j}label.grid(row=j, column=3, padx=15, pady=10)')
        exec(f'bonus{j}label.grid(row=j, column=4, padx=15, pady=10)')
        exec(f'totalsalary{j}label.grid(row=j, column=5, padx=20, pady=10)')

    # Add the frame to the canvas
    canvas.create_window(10, 10, window=frame, anchor="nw")

    # Update the scrollregion of the canvas
    root1.update()
    canvas.configure(scrollregion = canvas.bbox('all'))

    # Place the canvas and the scrollbar in the window
    canvas.place(x=20, y=200)
    scroll.place(x=860, y=200, height=355)

    addstaffbutton = Button(root1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                      text="Add New Staff", font=('Garamond', 20), bd=0, command = lambda:(root1.destroy(), addstaff(direct, datadirect)), cursor='hand2')
    addstaffbutton.place(x=150, y=570)    

    modifydetailsbutton = Button(root1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                      text="Modify Details", font=('Garamond', 20), bd=0, command = lambda:(root1.destroy(), modifydetails(direct, datadirect)), cursor='hand2')
    modifydetailsbutton.place(x=350, y=570)

    removestaffbutton = Button(root1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                      text="Remove Staff", font=('Garamond', 20), bd=0, command = lambda:(root1.destroy(), removestaff(direct, datadirect)), cursor='hand2')
    removestaffbutton.place(x=550, y=570)
    
    root1.mainloop()
