from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import customtkinter
import os
from openpyxl import *
from PIL import ImageTk


class Table:
    def __init__(self, name, capacity):
        self.name = "Table " + name
        self.capacity = capacity


column_names = {"Table Number": "A", "Capacity": "B"}
final_directory = ""

def excel(registration_number):
    global final_directory
    current_directory = os.getcwd()
    directory = registration_number
    final_directory = os.path.join(current_directory, directory)
    if not os.path.exists(final_directory):
        os.makedirs(final_directory)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tables"
    worksheet.append(list(column_names.keys()))
    final_directory += r"\Customer Service.xlsx"
    workbook.save(final_directory)


def working_hours(registration_number):
    def times():
        l = []
        for i in range(0,24):
            if len(str(i)) == 1:
                l.append(f'0{i}:00')
                l.append(f'0{i}:30')
            else:
                l.append(f'{i}:00')
                l.append(f'{i}:30')
        return l

    def check(event):
        if closingtimedropdown.get() != " " and len(closingtimedropdown.get())!=0:
            if (int(closingtimedropdown.get()[0:2]) < int(openingtimedropdown.get()[0:2])) or (int(closingtimedropdown.get()[0:2]) == int(openingtimedropdown.get()[0:2]) and int(closingtimedropdown.get()[3:]) <= int(openingtimedropdown.get()[3:])):
                closingtimedropdown["textvariable"] = " "
                    
    def closingtimes(event):
        x = times()
        l1=[]
        for i in x:
            if (int(i[0:2]) > int(openingtimedropdown.get()[0:2])):
                l1.append(i)
            elif int(i[0:2]) == int(openingtimedropdown.get()[0:2]) and int(i[3:]) > int(openingtimedropdown.get()[3:]):
                l1.append(i)
        closingtimedropdown["values"] = l1

    def addtiming():
        workbook = load_workbook("User_Data.xlsx")
        worksheet = workbook["User Data"]
        worksheet['I1'] = "Opening Time"
        worksheet['J1'] = "Closing Time"
        if (int(closingtimedropdown.get()[0:2]) > int(openingtimedropdown.get()[0:2])) or (int(closingtimedropdown.get()[0:2]) == int(openingtimedropdown.get()[0:2]) and int(closingtimedropdown.get()[3:]) > int(openingtimedropdown.get()[3:])):
             worksheet[f'I{worksheet.max_row}'] = openingtimedropdown.get()
             worksheet[f'J{worksheet.max_row}'] = closingtimedropdown.get()
        workbook.save("User_Data.xlsx")
        
    customtkinter.set_appearance_mode("dark")
    root = Tk()
    root.geometry('1366x768')
    root.state('zoomed')

    bgimage = ImageTk.PhotoImage(file='workinghours.png')
    bglabel = Label(root, image=bgimage)
    bglabel.place(x=0, y=0)

    openingtimelabel = Label(root, bg="white", fg="black", text="Opening Time", font=('Garamond', 22), bd=0)
    openingtimelabel.place(x=450, y=293)
 
    openingtimedropdown = ttk.Combobox(root, width=6, font=('Garamond', 22), textvariable=StringVar(value=" "))
    openingtimedropdown['values'] = times()
    openingtimedropdown.bind("<<ComboboxSelected>>", check)
    openingtimedropdown['state'] = 'readonly'
    openingtimedropdown.place(x=750, y=290)

    closingtimelabel = Label(root, bg="white", fg="black", text="Closing Time", font=('Garamond', 22), bd=0)
    closingtimelabel.place(x=450, y=357)

    closingtimedropdown = ttk.Combobox(root, width=6, font=('Garamond', 22), textvariable=StringVar(value=" "))
    closingtimedropdown["values"] = [""]
    closingtimedropdown.bind("<ButtonPress>", closingtimes)
    closingtimedropdown['state']='readonly'
    closingtimedropdown.place(x=750, y=350)

    continuebutton = Button(root, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                            text="Continue", font=('Garamond', 22), bd=0, cursor='hand2', command = lambda:(addtiming(), number_of_tables(registration_number, root, openingtimedropdown.get(), closingtimedropdown.get())))
    continuebutton.place(x=595, y=450)
    root.mainloop()

def number_of_tables(registration_number, root1, openingtime, closingtime):

    def tablecount_check():
        if (tablecount.get().strip().isdigit() is True and int(tablecount.get().strip()) > 0):
            tables(tablecount.get().strip(), root, openingtime, closingtime)
        else:
            messagebox.showerror('Error', 'Invalid Table Count')
    root1.destroy()
    excel(registration_number)
    customtkinter.set_appearance_mode("dark")
    root = Tk()
    root.geometry('1366x768')
    root.state('zoomed')

    bgimage = ImageTk.PhotoImage(file='tabledetails.png')
    bglabel = Label(root, image=bgimage)
    bglabel.place(x=0, y=0)

    tablecountlabel = Label(root, bg="white", fg="black", text="Number of Tables", font=('Garamond', 20), bd=0)
    tablecountlabel.place(x=555, y=300)

    tablecount = Entry(root, bg="white", fg="black", font=('Garamond', 20), width=5, bd=0, highlightthickness=1, justify="center")
    tablecount.configure(highlightbackground="black", highlightcolor="black")
    tablecount.place(x=615, y=357)

    continuebutton = Button(root, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                            text="Continue", font=('Garamond', 20), bd=0, cursor='hand2',
                            command=tablecount_check)
    continuebutton.place(x=590, y=425)
    root.mainloop()


def tables(tablecount, root1, openingtime, closingtime):

    def tabledata(capacity):
        newrow = {"SNo": "A", "Dish": "B", "Quantity": "C", "Rate": "D", "Amount": "E"}
        workbook = load_workbook(final_directory)
        worksheet = workbook["Tables"]
        for x in range(1, int(tablecount)+1):
            if capacity[x-1].get().strip() == "":
                messagebox.showerror('Error', f'Enter Capacity of Table {x}')
                break
            elif capacity[x-1].get().strip().isdigit() is False or int(capacity[x-1].get().strip()) < 0:
                messagebox.showerror('Error', f'Invalid Input for Table {x}')
                break
            else:
                exec(f'table_{x} = Table(str({x}),capacity[{x-1}].get())')
                exec(f'table = table_{x}')
                exec(f"worksheet[f'{column_names['''Table Number''']}{x+1}'] = table_{x}.name")
                exec(f"worksheet[f'{column_names['''Capacity''']}{x+1}'] = table_{x}.capacity")
                try:
                    worksheetnew = workbook[f'Table {x}']
                except KeyError:
                    workbook.create_sheet(title=f'Table {x}')
                    worksheetnew = workbook[f'Table {x}']
                    worksheetnew.append(list(newrow.keys()))
                workbook.save(final_directory)

    def timing(openingtime, closingtime):
        workbook = load_workbook(final_directory)
        worksheet = workbook["Tables"]
        l=[]
        for x in range(int(openingtime[0:2]), int(closingtime[0:2])+1):
            if x == int(openingtime[0:2]) and openingtime.endswith("30") is True:
                l.append(openingtime)
            elif x == int(closingtime[0:2]) and closingtime.endswith("00") is True:
                l.append(closingtime)
                break
            elif len(str(x)) == 1:
                l.append(f'0{x}:00')
                l.append(f'0{x}:30')
            else:
                l.append(f'{x}:00')
                l.append(f'{x}:30')

        d={}
        for i in l:
            if (67 + l.index(i)) > 90:
                d[i] = 'A'+chr(l.index(i)-24+65)
                continue
            d[i] = chr(67 + l.index(i))

        for i in d:
            for j in range(int(tablecount)+1):
                if j == 0:
                    worksheet[d[i]+str(j+1)] = i
                else:
                    worksheet[d[i]+str(j+1)] = False

        workbook.save(final_directory)
            


    root1.destroy()
    customtkinter.set_appearance_mode("dark")
    root = Tk()
    root.geometry('1366x768')
    root.state('zoomed')

    bgimage = ImageTk.PhotoImage(file='tabledetails.png')
    bglabel = Label(root, image=bgimage)
    bglabel.place(x=0, y=0)

    tablecountlabel = Label(root, bg="white", fg="black", text="Capacity", font=('Garamond', 25), bd=0)
    tablecountlabel.place(x=590, y=250)
    
    # Create the canvas widget
    canvas = Canvas(root, width=400, height=245, bg="white", bd=0, highlightbackground="white", highlightcolor="white")

    # Create the scrollbar widget
    scroll = Scrollbar(root, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
    canvas.config(yscrollcommand=scroll.set)

    # Create the frame widget
    frame = Frame(canvas, width=390, height=345, bg="white")

    capacity = []
    # Create labels and entry boxes and add them to the frame
    for i in range(1, int(tablecount)+1):
        exec(f'table{i}label = Label(frame, text=f"Table {i}", bg="white", fg="black", font=("Garamond", 20), bd=0)')
        exec(f'table{i}capacity = Entry(frame, bg="white", fg="black", font=("Garamond", 20), width=5, bd=0, highlightthickness=1, justify="center")')
        exec(f'table{i}capacity.configure(highlightbackground="black", highlightcolor="black")')
        exec(f'capacity.append(table{i}capacity)')
        exec(f'table{i}label.grid(row=i, column=0, padx=40, pady=8)')
        exec(f'table{i}capacity.grid(row=i, column=1, padx=90, pady=8)')

    # Add the frame to the canvas
    canvas.create_window(0, 0, window=frame, anchor="nw")

    # Update the scrollregion of the canvas
    root.update()
    canvas.configure(scrollregion = canvas.bbox('all'))

    # Place the canvas and the scrollbar in the window
    canvas.place(x=450, y=300)
    scroll.place(x=835, y=300, height=245)

    continuebutton = Button(root, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                            text="Continue", font=('Garamond', 20), bd=0, cursor='hand2',
                            command=lambda: (tabledata(capacity),timing(openingtime,closingtime)))
    continuebutton.place(x=590, y=570)

    # Start the main loop
    root.mainloop()

working_hours("11111111111111")


