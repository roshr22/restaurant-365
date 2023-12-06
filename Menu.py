from openpyxl import *
from tkinter import *
from tkinter import messagebox

col_names={"Dish name":"A","Price":"B"}

try:
    workbook = load_workbook("test.xlsx")
    sheet=workbook["Menu"]

except KeyError:
    workbook.create_sheet(title="Menu")
    sheet=workbook["Menu"]
    sheet.append(list(col_names.keys()))
    workbook.save("test.xlsx")

def update_prices():
    global Dishname, Price
    temp=0
    dish_name=Dishname
    P=Price
    for i in range(2, sheet.max_row + 1):
        Dname = sheet[f'{col_names["Dish name"]}{i}'].value
        pr=sheet[f'{col_names["Price"]}{i}'].value
        if Dname==dish_name and pr==P:
            messagebox.showerror("Error",f"The price of {dish_name} is already set to {P}")
            break
        elif Dname==dish_name and pr!=P:
            sheet[f'{col_names["Price"]}{i}']=P
            break
        workbook.save("test.xlsx")
            
    else:
        messagebox.showerror("Error",f"{dish_name} dosen't exist!")

def rem_dish():
    global Dishname
    temp=0
    dish_name=Dishname
    for i in range(2, sheet.max_row + 1):
        Dname = sheet[f'{col_names["Dish name"]}{i}'].value
        if Dname==dish_name:
            sheet[f'{col_names["Dish name"]}{i}']=''
            sheet[f'{col_names["Price"]}{i}']=''
            temp=i
            sheet.delete_rows(temp, 1)
            workbook.save("test.xlsx")
            break     
    else:
       messagebox.showerror("Error",f"{dish_name} doesn't exist")
def add_dish():
    global Dishname, Price
    dish_name=Dishname
    p=Price
    for i in range(2, sheet.max_row + 1):
        Dname = sheet[f'{col_names["Dish name"]}{i}'].value
        if Dname==dish_name:
            root2=Tk()
            root2.title("Update existing dish suggestion")
            f=Frame(root2,height=600,width=1000)
            f.pack()
            promptlabel=Label(f,text=f"{dish_name} exists! do you want to update prices?")
            promptlabel.place(x=300,y=200)
            yes=Button(f,text="YES",command=lambda:(update_prices(), close_window(root2)))
            yes.place(x=300,y=300)
            no=Button(f,text="NO",command=lambda:(close_window(root2)))
            no.place(x=400,y=300)
            root2.mainloop()               
    else:
        sheet[f'{col_names["Dish name"]}{sheet.max_row + 1}'] =dish_name
        sheet[f'{col_names["Price"]}{sheet.max_row}'] =p
        workbook.save("test.xlsx")
def delete(x):
    x.delete(0,'end')
        
def close_window(root1):
    root1.destroy()

def new_window1():
    global Dishname, Price
    def prod1():
        global Dishname, Price
        Dishname+=dish.get()
        Price+=int(price.get())
    root=Tk()
    root.title("Update existing menu")
    root.geometry("1200x800")
    f=Frame(root,height=600,width=1000)
    f.pack()
    dish=Entry(f, width=50)
    dish.place(x=300, y=300)
    dishlabel=Label(f, text="Dish name")
    dishlabel.place(x=100, y=300)
    price = Entry(f, width=50)
    price.place(x=300, y=350)
    pricelabel=Label(f, text="Price")
    pricelabel.place(x=100, y=350)
    add=Button(f,text="Add new dish",command=lambda:(prod1(),add_dish()))
    add.place(x=100,y=400)

Dishname=""
Price=0
newprice=0
root = Tk()
root.title("Update Menu")
root.geometry("1200x800")
fr=Frame(root,height=600,width=1000)
fr.pack()
Update_menu=Button(fr,text="Add dish", command=lambda:(close_window(root),new_window1()))
Update_menu.place(x=300,y=300)
remove_dish=Button(fr,text="Remove existing dish", command=lambda:(close_window(root),new_window2()))
remove_dish.place(x=600,y=300)

'''update=Button(f,text="Update prices",command=lambda:(prod1(),update_prices()))
    update.place(x=300,y=400)'''
'''def new_window1():
    global Dishname, Price
    def prod1():
        global Dishname, Price
        Dishname+=dish.get()
        Price+=int(price.get())
    
    root=Tk()
    root.title("Add new dish")
    root.geometry("1200x800")
    f=Frame(root,height=600,width=1000)
    f.pack()
    dish=Entry(f, width=50)
    dish.place(x=300, y=300)
    dishlabel=Label(f, text="Dish name")
    dishlabel.place(x=100, y=300)
    price = Entry(f, width=50)
    price.place(x=300, y=350)
    pricelabel=Label(f, text="Price")
    pricelabel.place(x=100, y=350)
    add=Button(f,text="Add", command=lambda:(prod1(),add_dish()))
    add.place(x=300,y=450)
    root.mainloop()

def new_window2():
    global Dishname
    def prod2():
        global Dishname
        Dishname+=dish.get()
    r=Tk()
    r.title("Remove existing dish")
    r.geometry("1200x800")
    f1=Frame(r,height=600,width=1000)
    f1.pack()
    dish=Entry(f1, width=50)
    dish.place(x=300, y=300)
    dishlabel=Label(f1, text="Dish name")
    dishlabel.place(x=100, y=300)
    remove=Button(f1,text="Remove", command=lambda:(prod2(),rem_dish()))
    remove.place(x=300,y=450)
    r.mainloop()

def new_window3():
    global Dishname, Price
    def prod2():
        global Dishname, Price
        Dishname+=dish.get()
        Price+=int(newprice.get())
    r1=Tk()
    r1.title("Update prices")
    r1.geometry("1200x800")
    f2=Frame(r1,height=600,width=1000)
    f2.pack()
    dish=Entry(f2, width=50)
    dish.place(x=300, y=300)
    dishlabel=Label(f2, text="Dish name")
    dishlabel.place(x=100, y=300)
    newprice = Entry(f2, width=50)
    newprice.place(x=300, y=350)
    newpricelabel=Label(f2, text="New Price")
    newpricelabel.place(x=100, y=350)
    update=Button(f2,text="Update prices", command=lambda:(prod2(),update_prices()))
    update.place(x=300,y=450)
    r1.mainloop()'''


'''Add=Button(fr,text="Add new dish", command=lambda:(close_window(),new_window1()))
Add.place(x=400,y=400)
Rem=Button(fr,text="Remove existing dish", command=lambda:(close_window(),new_window2()))
Rem.place(x=400,y=450)
Update=Button(fr,text="Update prices", command=lambda:(close_window(),new_window3()))
Update.place(x=400,y=500)
root.mainloop()'''
