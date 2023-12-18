from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import customtkinter
import bcrypt
import os
from openpyxl import *

#Signup Page
def signup_page(direct, root1):
    root1.destroy()
        
    def login(direct, root1):
        import loginpage
        loginpage.login_page(direct, root1)
    
    def restname(event):
        if restaurantname.get() == 'Enter Restaurant Name':
            restaurantname.delete(0, END)

    def regnum(event):
        if registrationnumber.get() == 'Enter Registration Number':
            registrationnumber.delete(0, END)

    def phoneno(event):
        if phonenumber.get() == 'Enter Phone Number':
            phonenumber.delete(0, END)

    def email(event):
        if emailid.get() == 'Enter Email Id':
            emailid.delete(0, END)

    def passwd(event):
        if password.get() == 'Enter Password':
            password.delete(0, END)
            password.configure(show='*')

    def addr(event):
        if address.get() == 'Enter Address':
            address.delete(0, END)

    def pincd(event):
        if pincode.get() == 'Enter Pincode':
            pincode.delete(0, END)

    def hide():
        closeeyeloc = os.path.join(direct, r'images\closeye.png')
        openeye.configure(file=closeeyeloc)
        password.configure(show='')
        eyebutton.configure(command=show)

    def show():
        openeyeloc = os.path.join(direct, r'images\openeye.png')
        openeye.configure(file=openeyeloc)
        password.configure(show='*')
        eyebutton.configure(command=hide)

    #Validating Input Data
    def adddata(direct):
        restaurant_name = restaurantname.get()
        registration_number = registrationnumber.get().strip()
        phone_number = phonenumber.get()
        email_id = emailid.get()
        password_ = password.get()
        address_ = address.get()
        state_ = state.get()
        pincode_ = pincode.get()
        if restaurant_name == "Enter Restaurant Name" or restaurant_name.strip() == "":
            messagebox.showerror('Error', 'Enter Restaurant Name')
        elif registration_number == "Enter Registration Number" or registration_number == "":
            messagebox.showerror('Error', 'Enter Registration Number')
        elif len(registration_number) != 14:
            messagebox.showerror('Error', 'Enter Valid Registration Number')
        elif registration_number[0] not in ("1", "2"):
            messagebox.showerror('Error', 'Enter Valid Registration Number')
        elif registration_number[1:3].isdigit() is False or (int(registration_number[1:3]) not in range(0, 36)):
            messagebox.showerror('Error', 'Enter Valid Registration Number')
        elif phone_number == "Enter Phone Number" or phone_number == "":
            messagebox.showerror('Error', 'Enter Phone Number')
        elif (len(phone_number) != 10 or phone_number.isdigit() is False
              or int(phone_number) not in range(7000000000, 10000000000)):
            messagebox.showerror('Error', 'Enter Valid Phone Number')
        elif email_id == "Enter Email Id" or email_id.strip() == "":
            messagebox.showerror('Error', 'Enter Email Id')
        elif (email_id.endswith("@gmail.com") is False and email_id.endswith("@yahoo.com") is False
              and email_id.endswith("@outlook.com") is False and email_id.endswith("@hotmail.com") is False
              and email_id.endswith("@icloud.com") is False):
            messagebox.showerror('Error', 'Enter Valid Email Id')
        elif password_ == "Enter Password" or password_ == "":
            messagebox.showerror('Error', 'Enter Password')
        elif (len(password_) < 7 or password_.isalpha() is True or password_.isdigit() is True
              or password_.isalnum() is True or password_.isupper() is True or password_.islower() is True
              or any(char.isdigit() for char in password_) is False):
            messagebox.showerror('Error',
                                 'Password must contain atleast 7 digits, 1 uppercase character, '
                                 '1 lowercase character, 1 digit and/or 1 special character')
        elif address_ == "Enter Address" or address_ == "":
            messagebox.showerror('Error', 'Enter Address')
        elif state_ not in state['values']:
            messagebox.showerror('Error', 'Choose State from the dropdown')
        elif pincode_.isdigit() is False or len(pincode_) != 6 or int(pincode_[0:2]) not in (11, 12, 13, 14, 15, 16, 17,
                                                                                             18, 19, 20, 21, 22, 23, 24,
                                                                                             25, 26, 27, 28, 29, 30, 31,
                                                                                             32, 33, 34, 36, 37, 38, 39,
                                                                                             40, 41, 42, 43, 44, 45, 46,
                                                                                             47, 48, 49, 50, 51, 52, 53,
                                                                                             56, 57, 58, 59, 60, 61, 62,
                                                                                             63, 64, 67, 68, 69, 70, 71,
                                                                                             72, 73, 74, 75, 76, 77, 78,
                                                                                             79, 80, 81, 82, 83, 84, 85,
                                                                                             90, 91, 92, 93, 94, 95, 96,
                                                                                             97, 98, 99):
            messagebox.showerror('Error', 'Enter Valid Pincode')
            messagebox.showerror('Error', 'Enter Valid Pincode')
        elif var.get() == 0:  # Error if terms & condition box is not checked
            messagebox.showerror('Error', 'Accept our Terms and Conditions')
        else:
            workbookloc = os.path.join(direct + r"User_Data.xlsx")
            column_names = {"Registration Number": "A", "Email": "B", "Password": "C", "Name": "D", "Phone Number": "E","Address": "F", "State": "G", "Pincode": "H"}
            try:
                workbook = load_workbook(workbookloc)
                worksheet = workbook["User Data"]
                for i in range(2, worksheet.max_row + 1):
                    registration_excel = worksheet[f'{column_names["Registration Number"]}{i}'].value
                    if registration_excel == registration_number:
                        messagebox.showerror("Error", f"Account exists! Please login")
                        break
                else:    
                    for i in range(2, worksheet.max_row + 1):
                        email_excel = worksheet[f'{column_names["Email"]}{i}'].value
                        if email_excel == email_id:
                            messagebox.showerror("Error", f"Account exists! Please login or try again with a different email id")
                            break
                    else:
                        password_ = password_.encode('utf-8')
                        salt = bcrypt.gensalt()
                        password_ = bcrypt.hashpw(password_, salt)
                        verification(direct, root, email_id, registration_number, password_, restaurant_name, phone_number, address_,state_, pincode_)
            except FileNotFoundError:
                password_ = password_.encode('utf-8')
                salt = bcrypt.gensalt()
                password_ = bcrypt.hashpw(password_, salt)
                verification(direct, root, email_id, registration_number, password_, restaurant_name, phone_number, address_,state_, pincode_)
        
    def verification(direct, root, email_id, registration_number, password_, restaurant_name, phone_number, address_,state_, pincode_):
        import emailverification
        emailverification.verification_window(direct, root, email_id, registration_number, password_, restaurant_name, phone_number, address_,state_, pincode_)

    #Creating new window    
    customtkinter.set_appearance_mode("dark")
    root = Tk()
    root.geometry('1366x768')
    root.state('zoomed')

    # Placing a background colour
    bglabel = Label(root, bg="white", width=1366, height=768)
    bglabel.place(x=0, y=0)

    signuplabel = Label(root, bg="white", fg="black", text="SIGN UP", font=('Garamond', 60, "bold"), bd=0)
    signuplabel.place(x=500, y=20)

    # Restaurant Name
    restaurantnamelabel = Label(root, bg="white", fg="black", text="Restaurant Name", font=('Garamond', 23),
                                bd=0)
    restaurantnamelabel.place(x=160, y=130)
    restaurantname = Entry(root, bg="white", fg="grey", font=('Garamond', 17), width=28, bd=0)
    restaurantname.place(x=160, y=170)
    restaurantname.insert(0, "Enter Restaurant Name")
    restaurantname.bind('<FocusIn>', restname)
    line1 = Frame(root, height=2, width=375, bg="black")
    line1.place(x=155, y=200)

    # Phone Number
    phonenumberlabel = Label(root, bg="white", fg="black", text="Phone Number", font=('Garamond', 23), bd=0)
    phonenumberlabel.place(x=160, y=210)
    phonenumber = Entry(root, bg="white", fg="grey", font=('Garamond', 17), width=28, bd=0)
    phonenumber.place(x=160, y=250)
    phonenumber.insert(0, "Enter Phone Number")
    phonenumber.bind('<FocusIn>', phoneno)
    line3 = Frame(root, height=2, width=375, bg="black")
    line3.place(x=155, y=280)

    # Email
    emailidlabel = Label(root, bg="white", fg="black", text="Email Id", font=('Garamond', 23), bd=0)
    emailidlabel.place(x=160, y=290)
    emailid = Entry(root, bg="white", fg="grey", font=('Garamond', 17), width=28, bd=0)
    emailid.place(x=160, y=330)
    emailid.insert(0, "Enter Email Id")
    emailid.bind('<FocusIn>', email)
    line4 = Frame(root, height=2, width=375, bg="black")
    line4.place(x=155, y=360)

    # Password
    passwordlabel = Label(root, bg="white", fg="black", text="Password", font=('Garamond', 23), bd=0)
    passwordlabel.place(x=160, y=370)
    password = Entry(root, bg="white", fg="grey", font=('Garamond', 17), width=28, bd=0)
    password.place(x=160, y=410)
    password.insert(0, "Enter Password")
    password.bind('<FocusIn>', passwd)
    line5 = Frame(root, height=2, width=375, bg="black")
    line5.place(x=155, y=445)
    openeyeloc = os.path.join(direct, r'images\openeye.png')
    openeye = PhotoImage(file=openeyeloc)
    eyebutton = Button(root, image=openeye, bd=0, bg='white', activebackground='white', cursor='hand2',
                       command=hide)
    eyebutton.place(x=520, y=415)

    # Registration Number
    registrationnumberlabel = Label(root, bg="white", fg="black", text="Registration Number",
                                    font=('Garamond', 23), bd=0)
    registrationnumberlabel.place(x=760, y=130)
    registrationnumber = Entry(root, bg="white", fg="grey", font=('Garamond', 17), width=28, bd=0)
    registrationnumber.place(x=760, y=170)
    registrationnumber.insert(0, "Enter Registration Number")
    registrationnumber.bind('<FocusIn>', regnum)
    line2 = Frame(root, height=2, width=375, bg="black")
    line2.place(x=755, y=200)

    # Address
    addresslabel = Label(root, bg="white", fg="black", text="Address", font=('Garamond', 23), bd=0)
    addresslabel.place(x=760, y=210)
    address = Entry(root, bg="white", fg="grey", font=('Garamond', 17), width=28, bd=0)
    address.place(x=760, y=250)
    address.insert(0, "Enter Address")
    address.bind('<FocusIn>', addr)
    line6 = Frame(root, height=2, width=375, bg="black")
    line6.place(x=755, y=280)

    # State
    statelabel = Label(root, bg="white", fg="black", text="State", font=('Garamond', 23), bd=0)
    statelabel.place(x=760, y=290)
    n = StringVar(value="Enter State") 
    state = ttk.Combobox(root, width=28, font=('Garamond', 17), textvariable=n)
    state['values'] = ('Andaman and Nicobar Islands', 'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Chandigarh', 'Chhattisgarh', 'Dadra and Nagar Haveli and Daman and Diu', 'Delhi', 'Goa', 'Gujarat', 'Haryana', 'Himachal Pradesh', 'Jammu and Kashmir', 'Jharkhand', 'Karnataka', 'Kerala', 'Ladakh', 'Lakshadweep', 'Madhya Pradesh', 'Maharashtra', 'Manipur', 'Meghalaya', 'Mizoram', 'Nagaland', 'Odisha', 'Puducherry', 'Punjab', 'Rajasthan', 'Sikkim', 'Tamil Nadu', 'Telangana', 'Tripura', 'Uttar Pradesh', 'Uttarakhand', 'West Bengal')
    state['state']='readonly'
    state.place(x=755, y=330)

    # Pincode
    pincodelabel = Label(root, bg="white", fg="black", text="Pincode", font=('Garamond', 23), bd=0)
    pincodelabel.place(x=760, y=370)
    pincode = Entry(root, bg="white", fg="grey", font=('Garamond', 17), width=28, bd=0)
    pincode.place(x=760, y=410)
    pincode.insert(0, "Enter Pincode")
    pincode.bind('<FocusIn>', pincd)
    line8 = Frame(root, height=2, width=375, bg="black")
    line8.place(x=755, y=445)

    var = IntVar()
    check = Checkbutton(root, height=2, width=65, bg='white', cursor='hand2', variable=var, font=('Garamond', 15),
                        text='I agree to the Terms of Service and Privacy Policy of Restaurant 365 and Metamorphosis')
    check.place(x=135, y=470)

    #Buttons
    signupbutton = Button(root, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                          text="Sign Up", font=('Garamond', 20), bd=0, command=lambda:(adddata(direct)), cursor='hand2')
    signupbutton.place(x=600, y=530)
    
    loginbutton = Button(root, bg="white", fg="black", activebackground="white", activeforeground="#015450",
                         text="Already have an account? Click here to login", font=('Garamond', 20, "underline"), bd=0,
                         command=lambda: (login(direct, root)), cursor='hand2')
    loginbutton.place(x=420, y=580)

    root.mainloop()
