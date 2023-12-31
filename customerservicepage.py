from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import customtkinter
from PIL import ImageTk
import os
from openpyxl import *

def landingpg(root, direct, datadirect):
    root.destroy()
    import landingpage
    landingpage.landingpage(direct, datadirect)
    
#Orders page: Asks which table to place order in

#Table's Orders Page
def orderspage(direct, datadirect):
    global order
    global billing

    #Order for a particular table
    def order(direct, datadirect, i):
        def adddata(datadirect, i, quantity):
            column_names = {"SNo": "A", "Dish": "B", "Quantity": "C", "Rate": "D", "Amount": "E"}
            workbook = load_workbook(datadirect + r"\Customer Service.xlsx")
            worksheet = workbook[f'Table {i}']

            column = {"Item": "A", "Price":"B"}
            workbook1 = load_workbook(datadirect + r"\Menu.xlsx")
            worksheet1 = workbook1["Menu"]

            existingitems = []
            originalitems = []
            
            for l in range(2, worksheet.max_row + 1):
                data = worksheet[f'{column_names["Dish"]}{l}'].value
                existingitems.append(data)
            
            for k in range(len(quantity)):
                quant = 0
                quant = int(quantity[k].get())
                name = worksheet1[f'{column["Item"]}{k+2}'].value
                rate = worksheet1[f'{column["Price"]}{k+2}'].value

                if quant != 0:
                    if name not in existingitems:
                        worksheet[f'{column_names["SNo"]}{worksheet.max_row+1}'] = f'{worksheet.max_row}'
                        worksheet[f'{column_names["Dish"]}{worksheet.max_row}'] = name
                        worksheet[f'{column_names["Quantity"]}{worksheet.max_row}'] = quant
                        worksheet[f'{column_names["Rate"]}{worksheet.max_row}'] = rate
                        worksheet[f'{column_names["Amount"]}{worksheet.max_row}'] = (float(rate) * quant)
                        workbook.save(datadirect + r"\Customer Service.xlsx")
                    else:
                        worksheet[f'{column_names["Quantity"]}{existingitems.index(name)+2}'] = worksheet[f'{column_names["Quantity"]}{existingitems.index(name)+2}'].value + quant
                        worksheet[f'{column_names["Amount"]}{existingitems.index(name)+2}'] = worksheet[f'{column_names["Quantity"]}{existingitems.index(name)+2}'].value + (float(rate) * quant)
                        workbook.save(datadirect + r"\Customer Service.xlsx")
                
                workbook.save(datadirect + r"\Customer Service.xlsx")
        root.destroy()
        column = {"Item": "A", "Price":"B"}
        workbook1 = load_workbook(datadirect + r"\Menu.xlsx")
        worksheet1 = workbook1["Menu"]
        
        customtkinter.set_appearance_mode("dark")
        root1 = Tk()
        root1.geometry('1366x768')
        root1.state('zoomed')

        bgloc = os.path.join(direct, r'images\orderspage.png')
        bgimage1 = ImageTk.PhotoImage(file=bgloc)
        bglabel = Label(root1, image=bgimage1)
        bglabel.place(x=0, y=0)

        tablelabel = Label(root1, text=f"TABLE {i}", bg="white", fg="black", font=("Garamond", 30), bd=0)
        tablelabel.place(x=565, y=150)

        # Create the canvas widget
        canvas = Canvas(root1, width=700, height=355, bg="white", bd=0, highlightbackground="white", highlightcolor="white")

        # Create the scrollbar widget
        scroll = Scrollbar(root1, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
        canvas.config(yscrollcommand=scroll.set)

        # Create the frame widget
        frame = Frame(canvas, width=690, height=345, bg="white")

        capacity = []
        quantity = []
        # Create labels and entry boxes and add them to the frame
        for j in range(2, int(worksheet1.max_row) + 1):  
            exec(f'global quantity{i}')
            textvalue = worksheet1[f'{column["Item"]}{j}'].value.title()
            pricevalue = worksheet1[f'{column["Price"]}{j}'].value
            #quantityvalue = worksheet1[f'{column["Quantity"]}{i}'].value
            exec(f'food{j}label = Label(frame, text=textvalue, bg="white", fg="black", font=("Garamond", 25), bd=0)')
            exec(f'price{j}label = Label(frame, text=pricevalue, bg="white", fg="black", font=("Garamond", 25), bd=0)')
            exec(f'quantity{j} = Spinbox(frame, from_=0, to=20, bg="white", disabledbackground="white", readonlybackground = "white", disabledforeground="black", fg="black", font=("Garamond", 25), bd=0, width=2, cursor="hand2", state = "readonly")')
            exec(f'food{j}label.grid(row=j, column=0, padx=70, pady=10)')
            exec(f'price{j}label.grid(row=j, column=1, padx=15, pady=10)')
            exec(f'quantity{j}.grid(row=j, column=2, padx=70, pady=10)')
            exec(f'quantity.append(quantity{j})')

        # Add the frame to the canvas
        canvas.create_window(10, 10, window=frame, anchor="nw")

        # Update the scrollregion of the canvas
        root.update()
        canvas.configure(scrollregion = canvas.bbox('all'))

        # Place the canvas and the scrollbar in the window
        canvas.place(x=240, y=200)
        scroll.place(x=955, y=200, height=355)

        placeorderbutton = Button(root1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                          text="Place Order", font=('Garamond', 25), bd=0, command = lambda:(adddata(datadirect, i, quantity), root1.destroy(), orderspage(direct, datadirect)), cursor='hand2')
        placeorderbutton.place(x=560, y=570)
        
        root1.mainloop()

    def billing(direct, datadirect, i):
        def cleartable(datadirect, i):
            column_names = {"SNo": "A", "Dish": "B", "Quantity": "C", "Rate": "D", "Amount": "E"}
            workbook = load_workbook(datadirect + r"\Customer Service.xlsx")
            worksheet = workbook[f'Table {i}']

            for i in range(2, int(worksheet.max_row) + 1):
                worksheet[f'{column_names["SNo"]}{i}'] = None
                worksheet[f'{column_names["Dish"]}{i}'] = None
                worksheet[f'{column_names["Quantity"]}{i}'] = None
                worksheet[f'{column_names["Rate"]}{i}'] = None
                worksheet[f'{column_names["Amount"]}{i}'] = None
                workbook.save(datadirect + r"\Customer Service.xlsx")
                        
        column = {"SNo": "A", "Dish": "B", "Quantity": "C", "Rate": "D", "Amount": "E"}
        workbook = load_workbook(datadirect + r"\Customer Service.xlsx")
        worksheet = workbook[f'Table {i}']

        root.destroy()
        
        customtkinter.set_appearance_mode("dark")
        root1 = Tk()
        root1.geometry('1366x768')
        root1.state('zoomed')

        bgimg = os.path.join(direct, r'images\billingpage.png')
        bgimage1 = ImageTk.PhotoImage(file=bgimg)
        bglabel = Label(root1, image=bgimage1)
        bglabel.place(x=0, y=0)

        tablelabel = Label(root1, text=f"TABLE {i}", bg="white", fg="black", font=("Garamond", 30), bd=0)
        tablelabel.place(x=565, y=150)

        # Create the canvas widget
        canvas = Canvas(root1, width=640, height=205, bg="white", bd=0, highlightbackground="white", highlightcolor="white")

        # Create the scrollbar widget
        scroll = Scrollbar(root1, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
        canvas.config(yscrollcommand=scroll.set)

        # Create the frame widget
        frame = Frame(canvas, width=640, height=205, bg="white")

        capacity = []
        quantity = []
        total = 0
        # Create labels and entry boxes and add them to the frame
        for j in range(1, int(worksheet.max_row) + 1):  
            snovalue = worksheet[f'{column["SNo"]}{j}'].value
            dishvalue = worksheet[f'{column["Dish"]}{j}'].value.title()
            quantityvalue = worksheet[f'{column["Quantity"]}{j}'].value
            ratevalue = worksheet[f'{column["Rate"]}{j}'].value
            amountvalue = worksheet[f'{column["Amount"]}{j}'].value
            if j != 1:
                total += amountvalue
            exec(f'sno{j}label = Label(frame, text=snovalue, bg="white", fg="black", font=("Garamond", 20), bd=0)')
            exec(f'dish{j}label = Label(frame, text=dishvalue, bg="white", fg="black", width=16, wraplength=260, font=("Garamond", 20), bd=0)')
            exec(f'quantity{j}label = Label(frame, text=quantityvalue, bg="white", fg="black", font=("Garamond", 20), bd=0)')
            exec(f'rate{j}label = Label(frame, text=ratevalue, bg="white", fg="black", font=("Garamond", 20), bd=0)')
            exec(f'amount{j}label = Label(frame, text=amountvalue, bg="white", fg="black", font=("Garamond", 20), bd=0)')
            exec(f'sno{j}label.grid(row=j, column=0, padx=15, pady=10)')
            exec(f'dish{j}label.grid(row=j, column=1, padx=0, pady=10)')
            exec(f'quantity{j}label.grid(row=j, column=2, padx=15, pady=10)')
            exec(f'rate{j}label.grid(row=j, column=3, padx=15, pady=10)')
            exec(f'amount{j}label.grid(row=j, column=4, padx=15, pady=10)')

        # Add the frame to the canvas
        canvas.create_window(10, 10, window=frame, anchor="nw")

        # Update the scrollregion of the canvas
        root.update()
        canvas.configure(scrollregion = canvas.bbox('all'))

        # Place the canvas and the scrollbar in the window
        canvas.place(x=300, y=200)
        scroll.place(x=940, y=200, height=205)

        totalamountlabel = Label(root1, text=f"Net Total: {str(round(total,2))}", bg="white", fg="black", font=("Garamond", 20), bd=0)
        totalamountlabel.place(x=752, y=410)

        cgstlabel = Label(root1, text=f"CGST at 2.5%: {str(round(0.025 * total,2))}", bg="white", fg="black", font=("Garamond", 20), bd=0)
        cgstlabel.place(x=700, y=440)

        sgstlabel = Label(root1, text=f"SGST at 2.5%: {str(round(0.025 * total,2))}", bg="white", fg="black", font=("Garamond", 20), bd=0)
        sgstlabel.place(x=705, y=470)

        gstlabel = Label(root1, text=f"Net GST: {str(round(0.05 * total,2))}", bg="white", fg="black", font=("Garamond", 20), bd=0)
        gstlabel.place(x=758, y=500)

        amountpayablelabel = Label(root1, text=f"Amount Payable: {str(round(total + (0.05 * total),2))}", bg="white", fg="black", font=("Garamond", 20, "bold"), bd=0)
        amountpayablelabel.place(x=666, y=530)
        
        placeorderbutton = Button(root1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",
                          text="Paid", font=('Garamond', 25), bd=0, cursor='hand2', command=lambda:(cleartable(datadirect, i), root1.destroy(), orderspage(direct, datadirect)))
        placeorderbutton.place(x=560, y=580)
        
        root1.mainloop()
        
    workbook = load_workbook(datadirect + r"\Customer Service.xlsx")
    worksheet = workbook["Tables"]

    customtkinter.set_appearance_mode("dark")
    root = Tk()
    root.geometry('1366x768')
    root.state('zoomed')

    bgimg = os.path.join(direct, r'images\orderpage.png')
    bgimage = ImageTk.PhotoImage(file=bgimg)
    bglabel = Label(root, image=bgimage)
    bglabel.place(x=0, y=0)

    liloc = os.path.join(direct, r'images\logohomebutton.png')
    logoimage = ImageTk.PhotoImage(file=liloc)
    homebutton = Button(root, image = logoimage , fg='white', bg='white', activeforeground='white',activebackground='white', cursor='hand2', bd=0, command=lambda:(landingpg(root, direct, datadirect)))
    homebutton.place(x=15, y=15)
    
    # Create the canvas widget
    canvas = Canvas(root, width=700, height=415, bg="white", bd=0, highlightbackground="white", highlightcolor="white")

    # Create the scrollbar widget
    scroll = Scrollbar(root, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
    canvas.config(yscrollcommand=scroll.set)

    # Create the frame widget
    frame = Frame(canvas, width=690, height=405, bg="white")

    capacity = []
    # Create labels and entry boxes and add them to the frame
    for i in range(1, int(worksheet.max_row)):
        global direct1, datadirect1
        direct1 = direct
        datadirect1 = datadirect
        exec(f'global table{i}orderbutton')
        exec(f'table{i}label = Label(frame, text=f"TABLE {i}", bg="white", fg="black", font=("Garamond", 25), bd=0)')
        exec(f'table{i}orderbutton = Button(frame, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="ORDER", font=("Garamond", 20), bd=0, cursor="hand2", width=12, command = lambda: (order(direct1, datadirect1, {i},)))')
        exec(f'table{i}billingbutton = Button(frame, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="BILLING", font=("Garamond", 20), bd=0, cursor="hand2", width=12, command = lambda: (billing(direct1, datadirect1, {i})))')
        exec(f'table{i}label.grid(row=i, column=0, padx=70, pady=10)')
        exec(f'table{i}orderbutton.grid(row=i, column=1, padx=15, pady=10)')
        exec(f'table{i}billingbutton.grid(row=i, column=2, padx=15, pady=10)')
        
    # Add the frame to the canvas
    canvas.create_window(10, 10, window=frame, anchor="nw")

    # Update the scrollregion of the canvas
    root.update()
    canvas.configure(scrollregion = canvas.bbox('all'))

    # Place the canvas and the scrollbar in the window
    canvas.place(x=-30, y=185)
    scroll.place(x=670, y=185, height=415)

    root.mainloop()
