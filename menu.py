from openpyxl import *
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import customtkinter
import os
from PIL import ImageTk
final_directory = "C:/Users/Ramesh's/Desktop/Final Restaurant 365/restaurant-365/11111111111111"
col_names={"Dish name":"A","Price":"B"}


def landingpg(root, direct, datadirect):
    root.destroy()
    import landingpage
    landingpage.landingpage(direct, datadirect)
    
def update_prices(direct, datadirect):
    global Dishname, nprice
    temp=0
    dish_name=Dishname.upper()
    P=nprice
    if dish_name.replace(' ','').isalnum()==False:
        messagebox.showerror("Error",f"Invalid dish name!")
    elif P == "" or (all(x.isdigit is False and x != "." for x in P) or P.count(".") > 1) or float(P.strip()) == 0:
        messagebox.showerror("Error",f"Invalid Input! Price must be a positive decimal value!")
        
    else:
        for i in range(2, sheet.max_row + 1):
            Dname = sheet[f'{col_names["Dish name"]}{i}'].value
            pr=sheet[f'{col_names["Price"]}{i}'].value
            if Dname==dish_name and pr==float(P):
                messagebox.showerror("Error",f"The price of {dish_name} is already set to {P}")
                break
            elif Dname==dish_name and pr!=float(P):
                sheet[f'{col_names["Price"]}{i}']=float(P)
                workbook.save(datadirect + r"\menu.xlsx")
                messagebox.showinfo(message=f"The price of {dish_name} has been updated to {P}")
                break
                
        else:
            messagebox.showerror("Error",f"{dish_name} dosen't exist!")

def rem_dish(direct, datadirect):
    global Dishname
    temp=0
    dish_name=Dishname.upper()
    if dish_name.replace(' ','').isalnum()==False:
        messagebox.showerror("Error","Invalid Input")
    else:
        for i in range(2, sheet.max_row + 1):
            Dname = sheet[f'{col_names["Dish name"]}{i}'].value
            if Dname==dish_name:
                sheet[f'{col_names["Dish name"]}{i}']=None
                sheet[f'{col_names["Price"]}{i}']=None
                temp=i
                sheet.delete_rows(temp, 1)
                workbook.save(datadirect + r"\menu.xlsx")
                messagebox.showinfo(message=f"{dish_name} has been removed from menu")
                break     
        else:
            messagebox.showerror("Error",f"{dish_name} doesn't exist")
            
def add_dish(direct, datadirect):
    global Dishname, Price
    dish_name=Dishname.upper()
    p=Price
    if dish_name.replace(' ','').isalnum()==False:
        messagebox.showerror("Error","Invalid Input! Invalid Dish Name!")
    elif p == "" or (all(x.isdigit is False and x != "." for x in p) or p.count(".") > 1) or float(p.strip()) == 0:
        messagebox.showerror("Error","Invalid Input! Price must be a positive float value!")
    else:
        for i in range(2, sheet.max_row + 1):
            Dname = sheet[f'{col_names["Dish name"]}{i}'].value
            if Dname==dish_name:
                messagebox.showerror("Error",f"{dish_name} already exists!")
                break
        else:
            sheet[f'{col_names["Dish name"]}{sheet.max_row + 1}'] =dish_name
            sheet[f'{col_names["Price"]}{sheet.max_row}'] = float(p)
            workbook.save(datadirect + r"\menu.xlsx")
            messagebox.showinfo(message=f"{dish_name} has been added to menu")
def delete(x):
    x.delete(0,'end')
        
def close_window(root1):
    root1.destroy()

def new_window1(direct, datadirect):
    global Dishname, Price
    def prod1():
        global Dishname, Price
        Dishname=dish.get()
        Price=price.get()
    
    r1=Tk()
    r1.title("Add new dish")
    r1.geometry("1366x768")
    r1.state('zoomed')
    
    nw1loc = os.path.join(direct, r'images\new_window1.png')
    bgimage = ImageTk.PhotoImage(file=nw1loc)
    bglabel = Label(r1, image=bgimage)
    bglabel.place(x=0, y=0)
    
    dish=Entry(r1, bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    dish.place(x=565, y=330)
    dishlabel=Label(r1,bg="white", fg="black", text="Dish Name", font=('Garamond', 22), bd=0)
    dishlabel.place(x=400, y=330)
    price = Entry(r1, bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    price.place(x=565, y=390)
    pricelabel=Label(r1,bg="white", fg="black", text="Price", font=('Garamond', 22), bd=0)
    pricelabel.place(x=400, y=390)
    add=Button(r1,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="ADD",font=('Garamond', 30), width=10,bd=0, cursor='hand2', command=lambda:(prod1(),add_dish(direct, datadirect)))
    add.place(x=530,y=515)
    cbloc = os.path.join(direct, r'images\close.png')
    img = ImageTk.PhotoImage(file=cbloc)
    close=Button(r1, image = img, fg="white", activebackground="white", activeforeground="white", bd=0, cursor='hand2', command=lambda:(close_window(r1),main_window(direct, datadirect)))
    close.place(x=869,y=52)
    r1.mainloop()

def new_window2(direct, datadirect):
    global Dishname
    def prod2():
        global Dishname
        Dishname=dish.get()
    r2=Tk()
    r2.title("Remove existing dish")
    r2.geometry("1366x768")
    r2.state('zoomed')

    bgloc = os.path.join(direct, r'images\new_window2.png')
    bgimage = ImageTk.PhotoImage(file=bgloc)
    bglabel = Label(r2, image=bgimage)
    bglabel.place(x=0, y=0)
    
    dish=Entry(r2, bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    dish.place(x=560, y=360)
    dishlabel=Label(r2,bg="white", fg="black", text="Dish name",font=('Garamond', 22), bd=0)
    dishlabel.place(x=395, y=360)
    remove=Button(r2,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="REMOVE",font=('Garamond', 30), width=10,bd=0, cursor='hand2', command=lambda:(prod2(),rem_dish(direct, datadirect)))
    remove.place(x=520,y=515)
    cbloc = os.path.join(direct, r'images\close.png')
    img = ImageTk.PhotoImage(file=cbloc)
    close=Button(r2,bg="red", image = img, fg="white", activebackground="#015450", activeforeground="white", bd=0, cursor='hand2', command=lambda:(close_window(r2),main_window(direct, datadirect)))
    close.place(x=869,y=52)
    r2.mainloop()

def new_window3(direct, datadirect):
    global Dishname, nprice
    def prod2():
        global Dishname, nprice
        Dishname=dish.get()
        nprice=newprice.get()
    r3=Tk()
    r3.title("Update prices")
    r3.geometry("1366x768")
    r3.state('zoomed')

    bgloc = os.path.join(direct, r'images\new_window3.png')
    bgimage = ImageTk.PhotoImage(file=bgloc)
    bglabel = Label(r3, image=bgimage)
    bglabel.place(x=0, y=0)
    
    dish=Entry(r3,bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    dish.place(x=560, y=320)
    dishlabel=Label(r3,bg="white", fg="black", text="Dish Name",font=('Garamond', 22), bd=0)
    dishlabel.place(x=400, y=320)
    newprice = Entry(r3,bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    newprice.place(x=560, y=380)
    newpricelabel=Label(r3,bg="white", fg="black", text="New Price",font=('Garamond', 22), bd=0)
    newpricelabel.place(x=400, y=380)
    update=Button(r3,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="UPDATE",font=('Garamond', 25), width=10,bd=0, cursor='hand2', command=lambda:(prod2(),update_prices(direct, datadirect)))
    update.place(x=550,y=520)
    cbloc = os.path.join(direct, r'images\close.png')
    img = ImageTk.PhotoImage(file=cbloc)
    close=Button(r3, image = img, fg="white", activebackground="white", activeforeground="white", bd=0, cursor='hand2', command=lambda:(close_window(r3),main_window(direct, datadirect)))
    close.place(x=869,y=52)
    r3.mainloop()

Dishname=""
Price=0
nprice=0

def main_window(direct, datadirect):
    global workbook
    global sheet
    try:
        workbook = load_workbook(datadirect + r"\menu.xlsx")
        sheet=workbook["Menu"]

    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Menu"
        sheet.append(list(col_names.keys()))
        workbook.save(datadirect + r"\menu.xlsx")
        
    except KeyError:
        workbook.create_sheet(title="Menu")
        sheet=workbook["Menu"]
        sheet.append(list(col_names.keys()))
        workbook.save(datadirect + r"\menu.xlsx")
        
    root=Tk()
    root.geometry("1366x768")
    root.state('zoomed')

    mploc = os.path.join(direct, r'images\menu.png')
    bgimage = ImageTk.PhotoImage(file=mploc)
    bglabel = Label(root, image=bgimage)
    bglabel.place(x=0, y=0)

    liloc = os.path.join(direct, r'images\logohomebutton.png')
    logoimage = ImageTk.PhotoImage(file=liloc)
    homebutton = Button(root, image = logoimage , fg='white', bg='white', activeforeground='white',activebackground='white', cursor='hand2', bd=0, command=lambda:(landingpg(root, direct, datadirect)))
    homebutton.place(x=15, y=15)

    canvas = Canvas(root, width=490, height=295, bg="white", bd=0, highlightbackground="white", highlightcolor="white")
    scroll = Scrollbar(root, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
    canvas.config(yscrollcommand=scroll.set)
    frame = Frame(canvas, width=480, height=285, bg="white")
    for i in range(2,sheet.max_row + 1):
        Dname = sheet[f'{col_names["Dish name"]}{i}'].value.title()
        pr=sheet[f'{col_names["Price"]}{i}'].value
        exec(f'dish_namelabel=Label(frame,text=f"{Dname}",width=16, wraplength=260, bg="white", fg="black", font=("Garamond", 22), bd=0)')
        exec(f'pricelabel=Label(frame,text=f"{pr}",bg="white", fg="black", font=("Garamond", 22), bd=0)')
        exec(f'dish_namelabel.grid(row=i, column=0, padx=40, pady=8)')
        exec(f'pricelabel.grid(row=i, column=1, padx=25, pady=8)')

    canvas.create_window(0, 0, window=frame, anchor="nw")
    root.update()
    canvas.configure(scrollregion = canvas.bbox('all'))
    canvas.place(x=120, y=210)
    scroll.place(x=610, y=210, height=295)
    
    Add=Button(root,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="ADD DISH", font=('Garamond', 20)
               , bd=0, cursor='hand2', command=lambda:(close_window(root),new_window1(direct, datadirect)))
    Add.place(x=80,y=570)
    Rem=Button(root,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="REMOVE DISH", font=('Garamond', 20), bd=0, cursor='hand2', command=lambda:(close_window(root),new_window2(direct, datadirect)))
    Rem.place(x=300,y=570)
    Update=Button(root,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="MODIFY PRICE", font=('Garamond', 20), bd=0, cursor='hand2', command=lambda:(close_window(root),new_window3(direct, datadirect)))
    Update.place(x=560,y=570)
    root.mainloop()


