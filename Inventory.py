#Inventory: Getting Ingredient and Quantity input from user
from openpyxl import *
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import customtkinter
import os
from PIL import ImageTk

column_names = {"Product":"A", "Quantity":"B", "Max Capacity":"C"}

def landingpg(root, direct, datadirect):
    root.destroy()
    import landingpage
    landingpage.landingpage(direct, datadirect)

def below_threshold():
    for i in range(2,workbook_sheet.max_row+1):
        product= workbook_sheet[f'{column_names["Product"]}{i}'].value
        quantity = workbook_sheet[f'{column_names["Quantity"]}{i}'].value
        m_quantity=workbook_sheet[f'{column_names["Max Capacity"]}{i}'].value
        if quantity<=0.1*m_quantity:
            messagebox.showerror("Error",f"Alert! {product} is below threshold! Purchase soon!")
        
def no_notfication():
    t=0
    for i in range(2,workbook_sheet.max_row+1):
        product= workbook_sheet[f'{column_names["Product"]}{i}'].value
        quantity = workbook_sheet[f'{column_names["Quantity"]}{i}'].value
        m_quantity=workbook_sheet[f'{column_names["Max Capacity"]}{i}'].value
        if quantity<=0.1*m_quantity:
            t=1
    if t==0:
        messagebox.showinfo(message="No new notifications!")

def addprod(datadirect):
    global prod1,quant1,maxv1
    prod_name=prod1.upper()
    quant_val=quant1
    max_val=maxv1
    if prod_name.replace(' ','').isalnum()==False:
        messagebox.showerror("Error","Invalid Ingredient Name!")
    elif quant_val.strip().isdigit() is False or int(quant_val) < 1:
        messagebox.showerror("Error","Invalid Input! Quantity must be a positive integer")
    elif max_val.strip().isdigit() is False or int(max_val) < 1:
        messagebox.showerror("Error","Invalid Input! Max Capacity must be a positive integer")
    else:
        quant_val = int(quant_val)
        max_val = int(max_val)
        for i in range(2, workbook_sheet.max_row + 1):
            product = workbook_sheet[f'{column_names["Product"]}{i}'].value
            if product==prod_name:
                messagebox.showerror("Error",f"{prod_name} exists!")
                break
        else:
            workbook_sheet[f'{column_names["Product"]}{workbook_sheet.max_row + 1}'] = prod_name
            workbook_sheet[f'{column_names["Quantity"]}{workbook_sheet.max_row}'] = quant_val
            workbook_sheet[f'{column_names["Max Capacity"]}{workbook_sheet.max_row}'] = max_val
            workbook.save(datadirect + r"/inventory.xlsx")
            messagebox.showinfo(message=f"{prod_name} has been added to inventory")
        
def remprod(datadirect):
    global prod1
    temp=0
    prod_name=prod1.upper()
    if prod_name.replace(' ','').isalpha()==False:
        messagebox.showerror("Error",f"Invalid Input!")
    else:
        for i in range(2, workbook_sheet.max_row + 1):
            product = workbook_sheet[f'{column_names["Product"]}{i}'].value
            if product==prod_name:
                workbook_sheet[f'{column_names["Product"]}{i}']=None
                workbook_sheet[f'{column_names["Quantity"]}{i}']=None
                workbook_sheet[f'{column_names["Max Capacity"]}{i}']=None
                temp=i
                workbook_sheet.delete_rows(temp, 1)
                workbook.save(datadirect + r"/inventory.xlsx")
                messagebox.showinfo(message=f"{prod_name} has been removed from inventory")
                break    
        else:
           messagebox.showerror("Error",f"{prod_name} doesn't exist")
  
    
def saveprod(datadirect):
    global prod1,quant1
    prod_name=prod1.upper()
    quant_val=quant1
    if prod_name.replace(' ','').isalnum()==False:
        messagebox.showerror("Error",f"Invalid Input!")
    elif quant_val.strip().isdigit() is False or int(quant_val) == 0:
        messagebox.showerror("Error",f"Invalid Input! Quantity must be a postive number!")
    else:
        for i in range(2, workbook_sheet.max_row + 1):
            product = workbook_sheet[f'{column_names["Product"]}{i}'].value
            quantity = workbook_sheet[f'{column_names["Quantity"]}{i}'].value
            maxcap = workbook_sheet[f'{column_names["Max Capacity"]}{i}'].value
            quant_val = int(quant_val)
            if product == prod_name and quantity:
                if quant_val > (maxcap-quantity):
                    messagebox.showerror("Error",f"Invalid Input! Quantity exceeds max capacity!")
                    break
                else:
                    workbook_sheet[f'{column_names["Quantity"]}{i}']=int(workbook_sheet[f'{column_names["Quantity"]}{i}'].value)+quant_val
                    workbook.save(datadirect + r"/inventory.xlsx")
                    messagebox.showinfo(message=f"{quant_val} of {prod_name} has been added to inventory")
                    break
            elif product == prod_name and not quantity:
                workbook_sheet[f'{column_names["Quantity"]}{i}']=quant_val
                workbook.save(datadirect + r"inventory.xlsx")
                messagebox.showinfo(message=f"{quant_val} of {prod_name} has been added to inventory")
                break
        else:
            messagebox.showerror("Error",f"{prod_name} dosen't exist!")
        
def removequantity(datadirect):
    global prod1,quant1
    prod_name=prod1.upper()
    quant_val=quant1
    if prod_name.replace(' ','').isalnum()==False:
        messagebox.showerror("Error",f"Invalid Input!")
    elif quant_val.strip().isdigit() is False or int(quant_val) == 0:
        messagebox.showerror("Error",f"Invalid Input!Quantity must be a postive number!")
    else:
        for i in range(2, workbook_sheet.max_row + 1):
            product = workbook_sheet[f'{column_names["Product"]}{i}'].value
            quantity = workbook_sheet[f'{column_names["Quantity"]}{i}'].value
            quant_val = int(quant_val)
            if product == prod_name and quantity:
                if quantity - quant_val < 0:
                    messagebox.showerror("Error","Quantity exceeds existing stock!")
                    break
                else:
                    workbook_sheet[f'{column_names["Quantity"]}{i}'] = workbook_sheet[f'{column_names["Quantity"]}{i}'].value - quant_val
                    workbook.save(datadirect + r"/inventory.xlsx")
                    messagebox.showinfo(message=f"{quant_val} of {prod_name}s has been removed from inventory")
                break
        else:
            messagebox.showerror("Error",f"{prod_name} dosen't exist in inventory!")

def close_window(root1):
    root1.destroy()
def new_window1(direct, datadirect, workbook, workbook_sheet):
    global prod1, quant1, maxv1
    def prod2():
        global prod1, quant1, maxv1
        prod1=prod.get()
        quant1=quant.get()
        maxv1=maxv.get()
    
    r1=Tk()
    r1.title("Add new product")
    r1.geometry("1366x768")
    r1.state("zoomed")

    niloc = os.path.join(direct, r'images\addnewingredientpage.png')
    bgimage = ImageTk.PhotoImage(file=niloc)
    bglabel = Label(r1, image=bgimage)
    bglabel.place(x=0, y=0)

    prod=Entry(r1, bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    prod.place(x=560, y=300)
    prodlabel=Label(r1,bg="white", fg="black", text="Ingredient",font=('Garamond', 22), bd=0)
    prodlabel.place(x=390, y=300)
    quant = Entry(r1, bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    quant.place(x=560, y=360)
    quantlabel=Label(r1,bg="white", fg="black", text="Quantity",font=('Garamond', 22), bd=0)
    quantlabel.place(x=390, y=360)
    maxv=Entry(r1, bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    maxv.place(x=560,y=420)
    maxvlabel=Label(r1, bg="white", fg="black", text="Max Capacity",font=('Garamond', 22), bd=0)
    maxvlabel.place(x=390, y=420)
    add=Button(r1, height=1, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="ADD PRODUCT", font=('Garamond', 28),bd=0, cursor='hand2',command=lambda:(prod2(),addprod(datadirect)))
    add.place(x=495,y=519)
    cbloc = os.path.join(direct, r'images\close.png')
    img = ImageTk.PhotoImage(file=cbloc)
    close=Button(r1, image = img, fg="white", activebackground="white", activeforeground="white", bd=0, cursor='hand2', command=lambda:(close_window(r1),main_window(direct, datadirect)))
    close.place(x=869,y=52)
    r1.mainloop()

def new_window2(direct, datadirect, workbook, workbook_sheet):
    global prod1
    def prod3():
        global prod1
        prod1=prod.get()

    r2=Tk()
    r2.title("Remove existing product")
    r2.geometry("1366x768")
    r2.state("zoomed")

    reiimg = os.path.join(direct, r'images\removeexistingingredientpage.png')
    bgimage = ImageTk.PhotoImage(file=reiimg)
    bglabel = Label(r2, image=bgimage)
    bglabel.place(x=0, y=0)
    
    prod=Entry(r2,bg="white", fg="black", font=('Garamond', 20), width=25, bd=0, highlightthickness=1, justify="center")
    prod.place(x=560, y=360)
    prodlabel=Label(r2, bg="white", fg="black", text="Ingredient",font=('Garamond', 22), bd=0)
    prodlabel.place(x=395, y=360)
    remove=Button(r2,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text='Remove',font=('Garamond', 30),width=12,bd=0, cursor='hand2', command=lambda:(prod3(),remprod(datadirect)))
    remove.place(x=495,y=500)
    cbimg = os.path.join(direct, r'images\close.png')
    img = ImageTk.PhotoImage(file=cbimg)
    close=Button(r2, image = img, fg="white", activebackground="#015450", activeforeground="white", bd=0, cursor='hand2', command=lambda:(close_window(r2),main_window(direct, datadirect)))
    close.place(x=869,y=52)
    r2.mainloop()

def new_window3(direct, datadirect, workbook, workbook_sheet):
    global prod1, quant1
    def prod4():
        global prod1, quant1
        prod1=prod.get()
        quant1=quant.get()

    r3=Tk()
    r3.title("Modify quantity")
    r3.geometry("1366x768")
    r3.state("zoomed")

    mqloc = os.path.join(direct, r'images\modifyquantitypage.png')
    bgimage = ImageTk.PhotoImage(file=mqloc)
    bglabel = Label(r3, image=bgimage)
    bglabel.place(x=0, y=0)
    
    prod=Entry(r3,bg="white", fg="black", font=('Garamond', 22), width=20, bd=0, highlightthickness=1, justify="center")
    prod.place(x=580, y=300)
    prodlabel=Label(r3, bg="white", fg="black", text="Ingredient",font=('Garamond', 22), bd=0)
    prodlabel.place(x=400, y=300)
    quant = Entry(r3,bg="white", fg="black", font=('Garamond', 22), width=20, bd=0, highlightthickness=1, justify="center")
    quant.place(x=580, y=360)
    quantlabel=Label(r3, bg="white", fg="black", text="Quantity",font=('Garamond', 22), bd=0)
    quantlabel.place(x=400, y=360)
    add=Button(r3,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="Add",font=('Garamond', 25),width=12,bd=0, cursor='hand2',command=lambda:(prod4(),saveprod(datadirect)))
    add.place(x=530,y=430)
    remove=Button(r3,bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="Remove",font=('Garamond', 25),width=12,bd=0, cursor='hand2',command=lambda:(prod4(),removequantity(datadirect)))
    remove.place(x=530,y=520)
    cbloc = os.path.join(direct, r'images\close.png')
    img = ImageTk.PhotoImage(file=cbloc)
    close=Button(r3, image = img, fg="white", activebackground="#015450", activeforeground="white",bd=0, cursor='hand2', command=lambda:(close_window(r3),main_window(direct, datadirect)))
    close.place(x=869,y=52)
    r3.mainloop()
    
        
    
    

prod1=""
quant1=0
maxv1=0
def main_window(direct, datadirect):
    global workbook
    global workbook_sheet
    try:
        workbook = load_workbook(datadirect + r"/inventory.xlsx")
        workbook_sheet=workbook["Inventory"]

    except FileNotFoundError:
        workbook = Workbook()
        workbook_sheet = workbook.active
        workbook_sheet.title = "Inventory"
        workbook_sheet.append(list(column_names.keys()))
        workbook.save(datadirect + r"/inventory.xlsx")

    except KeyError:
        workbook
        workbook.create_sheet(title="Inventory")
        workbook_sheet=workbook["Inventory"]
        workbook_sheet.append(list(column_names.keys()))
        workbook.save(datadirect + r"/inventory.xlsx")
        
    root = Tk()
    root.title("Inventory")
    root.geometry("1366x768")
    root.state("zoomed")

    imgloc = os.path.join(direct, r'images\inventorypage.png')
    bgimage = ImageTk.PhotoImage(file=imgloc)
    bglabel = Label(root, image=bgimage)
    bglabel.place(x=0, y=0)

    liloc = os.path.join(direct, r'images\logohomebutton.png')
    logoimage = ImageTk.PhotoImage(file=liloc)
    homebutton = Button(root, image = logoimage , fg='white', bg='white', activeforeground='white',activebackground='white', cursor='hand2', bd=0, command=lambda:(landingpg(root, direct, datadirect)))
    homebutton.place(x=15, y=15)

    canvas = Canvas(root, width=540, height=305, bg="white", bd=0, highlightbackground="white", highlightcolor="white")
    scroll = Scrollbar(root, orient=VERTICAL, command=canvas.yview, bd=0, highlightbackground="white", highlightcolor="white", troughcolor="black")
    canvas.config(yscrollcommand=scroll.set)
    frame = Frame(canvas, width=530, height=295, bg="white")
    for i in range(2,workbook_sheet.max_row+1):
        Prod = workbook_sheet[f'{column_names["Product"]}{i}'].value
        Q=workbook_sheet[f'{column_names["Quantity"]}{i}'].value
        Mv=workbook_sheet[f'{column_names["Max Capacity"]}{i}'].value
        exec(f'prod_namelabel=Label(frame,text=f"{Prod}",bg="white", fg="black", font=("Garamond", 20), bd=0)')
        exec(f'quantlabel=Label(frame,text=f"{Q}",bg="white", fg="black", font=("Garamond", 20), bd=0)')
        exec(f'maxcapacitylabel=Label(frame,text=f"{Mv}",bg="white", fg="black", font=("Garamond", 20), bd=0)')
        exec(f'prod_namelabel.grid(row=i, column=0, padx=20, pady=8)')
        exec(f'quantlabel.grid(row=i, column=1, padx=70, pady=8)')
        exec(f'maxcapacitylabel.grid(row=i, column=2, padx=70, pady=8)')

    canvas.create_window(0, 0, window=frame, anchor="nw")
    root.update()
    canvas.configure(scrollregion = canvas.bbox('all'))
    canvas.place(x=175, y=240)
    scroll.place(x=745, y=240, height=305)

    ingredient_label=Label(root,bg="white", fg="black", text="INGREDIENT",font=('Garamond', 20), bd=0)
    ingredient_label.place(x=175,y=200)
    quantity_label=Label(root,bg="white", fg="black", text="QUANTITY",font=('Garamond', 20), bd=0)
    quantity_label.place(x=370,y=200)
    mcapacity_label=Label(root,bg="white", fg="black", text="MAX CAPACITY",font=('Garamond', 20), bd=0)
    mcapacity_label.place(x=540,y=200)
    add_newprod=Button(root, height=2, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="ADD INGREDIENT",font=('Garamond', 15), bd=0, cursor='hand2', command=lambda:(close_window(root),new_window1(direct, datadirect, workbook, workbook_sheet)))
    add_newprod.place(x=80,y=570)
    remove_prod=Button(root, height=2, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="REMOVE INGREDIENT",font=('Garamond', 15), bd=0, cursor='hand2', command=lambda:(close_window(root),new_window2(direct, datadirect, workbook, workbook_sheet)))
    remove_prod.place(x=330,y=570)
    modify=Button(root, height=2, bg="#015450", fg="white", activebackground="#015450", activeforeground="white",text="MODIFY QUANTITY",font=('Garamond', 15), bd=0, cursor='hand2',command=lambda:(close_window(root),new_window3(direct, datadirect, workbook, workbook_sheet)))
    modify.place(x=620,y=570)
    nbloc = os.path.join(direct, r'images\notification.png')
    img = ImageTk.PhotoImage(file=nbloc)
    notify=Button(root,bg="white",image=img, fg="white", activebackground="white", activeforeground="white", bd=0, cursor='hand2',command=lambda:(no_notfication(),below_threshold()))
    notify.place(x=820,y=20)
    root.mainloop()



